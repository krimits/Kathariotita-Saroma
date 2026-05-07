from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.worksheet.worksheet import Worksheet


DATA_START_ROW = 8
STREET_COLUMN = 2
PREVIOUS_WEEK_COLUMNS = (3, 4, 5)
DAY_COLUMNS = (6, 7, 8, 9, 10, 11, 12)
DATA_COLUMNS = PREVIOUS_WEEK_COLUMNS + DAY_COLUMNS


def canonical_data_column_for_duplicate_point(
    row_values: dict[int, int | None],
    point_id: int,
    *,
    value_counts: dict[int, int],
) -> int | None:
    """Όταν το ίδιο point_id εμφανίζεται σε >1 στήλες C–L στην ίδια γραμμή, επιστρέφει τη στήλη της **μίας** εγγραφής «συλλεχθέν».

    Έτσι αποφεύγεται το διπλό KPI συλλεχθέντων (πορτοκαλί + πράσινο με τον ίδιο αριθμό). Στις στήλες ημερών
    προτιμάται η **δεξιότερη** (συνήθως το πράσινο ταίριασμα)· αν η επανάληψη είναι μόνο σε στήλες προηγούμενων
    εβδομάδων, η μικρότερη στήλη στο DATA_COLUMNS.
    """
    if value_counts.get(point_id, 0) <= 1:
        return None
    dup_cols = [c for c in DATA_COLUMNS if row_values.get(c) == point_id]
    if len(dup_cols) <= 1:
        return None
    day_dups = [c for c in dup_cols if c in DAY_COLUMNS]
    return max(day_dups) if day_dups else min(dup_cols)

GREEN_RGB = {"FF92D050", "FFA8D08D", "FFE2EFD9"}
ORANGE_RGB = {"FFFFC000", "FFFF9900", "FFC2A218", "FFF9CB9C", "FFFBE4D5"}
RED_RGB = {"FFFF0000", "FFC53030"}

# Background canonical RGB (distance threshold aligns with google_sheets_extractor).
_BACKGROUND_CANONICALS: tuple[tuple[str, tuple[int, int, int]], ...] = (
    ("red", (255, 0, 0)),
    ("orange", (255, 192, 0)),
    ("orange_alt", (255, 153, 0)),
    ("yellow", (255, 255, 0)),
    ("green", (146, 208, 80)),
)
_BACKGROUND_DISTANCE_MAX = 90.0


@dataclass(frozen=True)
class DashboardRecord:
    workbook: str
    sheet: str
    week_label: str
    row: int
    column: str
    street: str
    street_normalized: str
    period: str
    day: str
    point_id: int
    status: str
    source: str
    color: str
    confidence: str
    note: str


@dataclass(frozen=True)
class DashboardSummary:
    workbook: str
    sheet: str
    week_label: str
    total_records: int
    pending: int
    collected: int
    unknown_status: int
    pending_by_source: dict[str, int]
    collected_by_day: dict[str, int]
    resolved_previous_week_records: int
    pending_by_day: dict[str, int]
    top_pending_streets: list[dict[str, Any]]
    week_status: str
    extraction_warnings: list[str]


def normalize_street(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = re.sub(r"\s+", " ", text)
    return text


def cell_rgb(cell: Cell) -> str:
    color = cell.fill.fgColor
    if color.type == "rgb" and color.rgb:
        return str(color.rgb).upper()
    return ""


def fill_hex_to_rgb_tuple(fill_hex: str) -> tuple[int, int, int] | None:
    if not fill_hex or fill_hex.upper() == "00000000":
        return None
    h = fill_hex.upper().lstrip("#")
    if len(h) == 8 and h.startswith("FF"):
        h = h[2:]
    if len(h) != 6:
        return None
    try:
        return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    except ValueError:
        return None


def is_effective_transparent_sheet_background(rgb: tuple[int, int, int] | None) -> bool:
    if rgb is None:
        return True
    r, g, b = rgb
    return min(r, g, b) >= 245


def classify_background_bucket(
    bg_rgb: tuple[int, int, int] | None,
    *,
    transparent_fill: bool,
) -> str:
    if transparent_fill or bg_rgb is None:
        return "transparent"
    best_tag = "unknown"
    best_distance = float("inf")
    for tag, canonical in _BACKGROUND_CANONICALS:
        distance = math.sqrt(sum((a - b) ** 2 for a, b in zip(bg_rgb, canonical)))
        if distance < best_distance:
            best_distance = distance
            best_tag = tag
    if best_distance > _BACKGROUND_DISTANCE_MAX:
        return "unknown"
    if best_tag in ("orange", "orange_alt", "yellow"):
        return "orange"
    return best_tag


def is_foreground_red(rgb: tuple[int, int, int]) -> bool:
    r, g, b = rgb
    return r >= 190 and g <= 95 and b <= 95


def is_foreground_dark(rgb: tuple[int, int, int]) -> bool:
    r, g, b = rgb
    luminance = 0.299 * r + 0.587 * g + 0.114 * b
    return luminance <= 140


def pending_day_source_from_rgb(
    bg_rgb: tuple[int, int, int] | None,
    fg_rgb: tuple[int, int, int] | None,
    *,
    transparent_fill: bool,
    font_theme_dark: bool,
) -> tuple[str, str]:
    bucket = classify_background_bucket(bg_rgb, transparent_fill=transparent_fill)

    if bucket == "red":
        return "veltio", "pending_day_red_fill"

    if bucket == "green":
        return "unknown", "pending_day_green_fill_unmatched"

    if bucket == "orange":
        if fg_rgb is not None and is_foreground_red(fg_rgb):
            return "veltio", "pending_day_orange_fill_red_font"
        if fg_rgb is not None and is_foreground_dark(fg_rgb):
            return "supervisor", "pending_day_orange_fill_dark_font"
        if font_theme_dark:
            return "supervisor", "pending_day_orange_fill_theme_dark_font"
        return "supervisor", "pending_day_orange_fill_default_supervisor"

    if bucket == "transparent":
        if fg_rgb is not None and is_foreground_red(fg_rgb):
            return "veltio", "pending_day_conditional_red_font"
        if font_theme_dark:
            return "supervisor", "pending_day_conditional_theme_dark_font"
        if fg_rgb is None:
            return "supervisor", "pending_day_transparent_default_supervisor"
        if is_foreground_dark(fg_rgb):
            return "supervisor", "pending_day_conditional_dark_font"
        return "unknown", "pending_day_transparent_unknown_font"

    if fg_rgb is not None and is_foreground_red(fg_rgb):
        return "veltio", "pending_day_unknown_fill_red_font"
    if font_theme_dark or (fg_rgb is not None and is_foreground_dark(fg_rgb)):
        return "supervisor", "pending_day_unknown_fill_dark_font"
    return "unknown", "pending_day_unknown_fill"


def cell_foreground_rgb(cell: Cell) -> tuple[int, int, int] | None:
    fc = cell.font.color
    if fc is None:
        return None
    if fc.type == "rgb" and fc.rgb:
        raw = str(fc.rgb).upper()
        if raw.startswith("FF") and len(raw) == 8:
            body = raw[2:]
        elif len(raw) == 6:
            body = raw
        else:
            return None
        try:
            return int(body[0:2], 16), int(body[2:4], 16), int(body[4:6], 16)
        except ValueError:
            return None
    return None


def cell_font_theme_dark(cell: Cell) -> bool:
    fc = cell.font.color
    if fc is None:
        return False
    return fc.type == "theme" and getattr(fc, "theme", None) == 1 and float(getattr(fc, "tint", 0.0)) == 0.0


def direct_color_status(rgb: str) -> tuple[str, str]:
    if rgb in GREEN_RGB:
        return "collected", "unknown"
    if rgb in RED_RGB:
        return "pending", "veltio"
    if rgb in ORANGE_RGB:
        return "pending", "supervisor"
    return "unknown", "unknown"


def is_numeric(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def period_name(ws: Worksheet, column: int) -> str:
    value = ws.cell(5, column).value or ws.cell(6, column).value or ""
    return str(value).strip()


def day_name(ws: Worksheet, column: int) -> str:
    if column in DAY_COLUMNS:
        return period_name(ws, column)
    return ""


def week_label(ws: Worksheet) -> str:
    value = ws["F2"].value
    return str(value).strip() if value else ""


def iter_conditional_fill_colors(
    conditional_formatting: ConditionalFormattingList,
) -> Iterable[str]:
    for cf_range in conditional_formatting:
        for rule in conditional_formatting[cf_range]:
            dxf = getattr(rule, "dxf", None)
            fill = getattr(dxf, "fill", None)
            fg_color = getattr(fill, "fgColor", None)
            rgb = getattr(fg_color, "rgb", None)
            if rgb:
                yield str(rgb).upper()


def workbook_has_conditional_colors(ws: Worksheet) -> bool:
    return any(iter_conditional_fill_colors(ws.conditional_formatting))


def infer_status_from_row(ws: Worksheet, row: int, column: int, point_id: int) -> tuple[str, str, str]:
    """Ταιριάζει στο google_sheets_extractor.classify_row_value: duplicate point_id στη γραμμή C:L ⇒ συλλεχθέν.

    Το extract_records εκπέμπει μόνο μία εγγραφή ανά (σειρά, point_id)· τα υπόλοιπα κελιά παραλείπονται.
    """
    same_value_columns = [
        current_column
        for current_column in DATA_COLUMNS
        if is_numeric(ws.cell(row, current_column).value)
        and int(ws.cell(row, current_column).value) == point_id
    ]

    if len(same_value_columns) > 1:
        if column in DAY_COLUMNS:
            return "collected", "unknown", "inferred_duplicate_day_value"
        return "collected", "unknown", "inferred_duplicate_previous_value"

    cell = ws.cell(row, column)
    fill_hex = cell_rgb(cell)

    if column in PREVIOUS_WEEK_COLUMNS:
        return "pending", "previous_weeks", "unmatched_previous_week_value"

    if column in DAY_COLUMNS:
        transparent_fill = not fill_hex or fill_hex.upper() == "00000000"
        bg_tuple = None if transparent_fill else fill_hex_to_rgb_tuple(fill_hex)
        fg_tuple = cell_foreground_rgb(cell)
        theme_dark = cell_font_theme_dark(cell)
        bucket = classify_background_bucket(bg_tuple, transparent_fill=transparent_fill)

        if bucket == "green":
            return "pending", "unknown", "unmatched_day_numeric_single_cell_green"

        source, confidence = pending_day_source_from_rgb(
            bg_tuple,
            fg_tuple,
            transparent_fill=transparent_fill,
            font_theme_dark=theme_dark,
        )
        return "pending", source, confidence

    return "unknown", "unknown", "unclassified_numeric_value"


def extract_records(workbook_path: Path) -> tuple[list[DashboardRecord], list[str]]:
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb.worksheets[0]
    warnings: list[str] = []

    if workbook_has_conditional_colors(ws):
        warnings.append(
            "The workbook uses conditional formatting. Local XLSX parsing can infer many statuses, "
            "but Google Sheets effective formats are needed for exact red/orange source attribution."
        )

    records: list[DashboardRecord] = []
    label = week_label(ws)

    for row in range(DATA_START_ROW, ws.max_row + 1):
        street = ws.cell(row, STREET_COLUMN).value
        normalized_street = normalize_street(street)
        if not normalized_street:
            continue

        row_values: dict[int, int | None] = {}
        for column in DATA_COLUMNS:
            cell = ws.cell(row, column)
            if is_numeric(cell.value):
                row_values[column] = int(cell.value)

        value_counts: dict[int, int] = {}
        for value in row_values.values():
            value_counts[value] = value_counts.get(value, 0) + 1

        for column in DATA_COLUMNS:
            cell = ws.cell(row, column)
            if not is_numeric(cell.value):
                continue

            point_id = int(cell.value)
            canon = canonical_data_column_for_duplicate_point(
                row_values, point_id, value_counts=value_counts
            )
            if canon is not None and column != canon:
                continue

            status, source, confidence = infer_status_from_row(ws, row, column, point_id)
            records.append(
                DashboardRecord(
                    workbook=workbook_path.name,
                    sheet=ws.title,
                    week_label=label,
                    row=row,
                    column=cell.coordinate,
                    street=str(street).strip(),
                    street_normalized=normalized_street,
                    period=period_name(ws, column),
                    day=day_name(ws, column),
                    point_id=point_id,
                    status=status,
                    source=source,
                    color=cell_rgb(cell),
                    confidence=confidence,
                    note="",
                )
            )

    return records, warnings


def build_summary(
    workbook_path: Path,
    records: list[DashboardRecord],
    warnings: list[str],
) -> DashboardSummary:
    sheet = records[0].sheet if records else ""
    label = records[0].week_label if records else ""

    pending_records = [record for record in records if record.status == "pending"]
    collected_records = [record for record in records if record.status == "collected"]
    unknown_records = [record for record in records if record.status == "unknown"]

    pending_by_source = Counter(record.source for record in pending_records)
    collected_by_day = Counter(record.day for record in collected_records if record.day)
    resolved_previous_week_records = sum(
        1
        for record in collected_records
        if not record.day and record.period.startswith("ΕΚΚΡΕΜΟΤΗΤΕΣ")
    )
    pending_by_day = Counter(record.day or record.period for record in pending_records)
    top_pending = Counter(record.street_normalized for record in pending_records).most_common(15)

    pending_count = len(pending_records)
    if pending_count <= 100:
        week_status = "excellent"
    elif pending_count < 300:
        week_status = "good"
    else:
        week_status = "red"

    return DashboardSummary(
        workbook=workbook_path.name,
        sheet=sheet,
        week_label=label,
        total_records=len(records),
        pending=pending_count,
        collected=len(collected_records),
        unknown_status=len(unknown_records),
        pending_by_source=dict(sorted(pending_by_source.items())),
        collected_by_day=dict(sorted(collected_by_day.items())),
        resolved_previous_week_records=resolved_previous_week_records,
        pending_by_day=dict(sorted(pending_by_day.items())),
        top_pending_streets=[
            {"street": street, "pending": count}
            for street, count in top_pending
        ],
        week_status=week_status,
        extraction_warnings=warnings,
    )


def write_records_csv(records: list[DashboardRecord], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(DashboardRecord.__dataclass_fields__.keys())
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(asdict(record))


def write_summary_json(summary: DashboardSummary, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(asdict(summary), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def default_workbook(folder: Path) -> Path:
    workbooks = sorted(
        path
        for path in folder.glob("*.xlsx")
        if not path.name.startswith("~$")
    )
    if not workbooks:
        raise FileNotFoundError(f"No .xlsx workbook found in {folder}")
    return workbooks[0]


def run(workbook_path: Path, output_dir: Path) -> DashboardSummary:
    records, warnings = extract_records(workbook_path)
    summary = build_summary(workbook_path, records, warnings)
    write_records_csv(records, output_dir / "dashboard_records.csv")
    write_summary_json(summary, output_dir / "dashboard_summary.json")
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract dashboard-ready data from bulky waste Excel workbook.")
    parser.add_argument("--workbook", type=Path, default=None, help="Path to the XLSX workbook.")
    parser.add_argument("--output-dir", type=Path, default=Path("output"), help="Directory for CSV/JSON exports.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook or default_workbook(Path.cwd())
    summary = run(workbook_path.resolve(), args.output_dir.resolve())
    print(json.dumps(asdict(summary), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
