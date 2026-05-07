import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from dashboard_extractor import (
    DAY_COLUMNS,
    DashboardRecord,
    build_summary,
    canonical_data_column_for_duplicate_point,
    direct_color_status,
    extract_records,
    infer_status_from_row,
    normalize_street,
)

GREEN_FILL = PatternFill(fill_type="solid", fgColor="FF92D050")
ORANGE_FILL = PatternFill(fill_type="solid", fgColor="FFFFC000")


class DashboardExtractorTests(unittest.TestCase):
    def test_canonical_duplicate_prefers_rightmost_day_column(self) -> None:
        row_values = {6: 12, 7: 12}
        counts = {12: 2}
        self.assertEqual(
            canonical_data_column_for_duplicate_point(row_values, 12, value_counts=counts),
            7,
        )

    def test_canonical_duplicate_previous_week_only_uses_min_column(self) -> None:
        row_values = {3: 9, 4: 9}
        counts = {9: 2}
        self.assertEqual(
            canonical_data_column_for_duplicate_point(row_values, 9, value_counts=counts),
            3,
        )

    def test_normalize_street_trims_uppercases_and_collapses_spaces(self) -> None:
        self.assertEqual(normalize_street("  Βασιλίσσης   Όλγας  "), "ΒΑΣΙΛΊΣΣΗΣ ΌΛΓΑΣ")

    def test_direct_red_fill_maps_to_pending_veltio(self) -> None:
        self.assertEqual(direct_color_status("FFFF0000"), ("pending", "veltio"))

    def test_direct_orange_fill_maps_to_pending_supervisor(self) -> None:
        self.assertEqual(direct_color_status("FFFFC000"), ("pending", "supervisor"))

    def test_direct_green_fill_maps_to_collected(self) -> None:
        self.assertEqual(direct_color_status("FF92D050"), ("collected", "unknown"))

    def test_unknown_fill_stays_unknown(self) -> None:
        self.assertEqual(direct_color_status("00000000"), ("unknown", "unknown"))

    def test_infer_duplicate_before_color_green_is_collected(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(8, 2, "ΟΔΟΣ ΔΟΚΙΜΗΣ")
        ws.cell(8, 5, 42)
        ws.cell(8, 6, 42)
        ws.cell(8, 6).fill = GREEN_FILL
        ws.cell(5, 6, "ΔΕΥΤΕΡΑ")

        status, source, confidence = infer_status_from_row(ws, 8, 6, 42)
        self.assertEqual((status, source), ("collected", "unknown"))
        self.assertEqual(confidence, "inferred_duplicate_day_value")

    def test_infer_unique_green_day_cell_is_pending_not_collected(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(8, 2, "ΟΔΟΣ ΔΟΚΙΜΗΣ")
        ws.cell(8, 6, 7)
        ws.cell(8, 6).fill = GREEN_FILL
        ws.cell(5, 6, "ΤΡΙΤΗ")

        status, source, confidence = infer_status_from_row(ws, 8, 6, 7)
        self.assertEqual(status, "pending")
        self.assertEqual(source, "unknown")
        self.assertEqual(confidence, "unmatched_day_numeric_single_cell_green")

    def test_infer_unique_previous_week_column_is_pending_previous_weeks(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(8, 2, "ΟΔΟΣ ΔΟΚΙΜΗΣ")
        ws.cell(8, 5, 99)
        ws.cell(5, 5, "ΕΚΚΡΕΜΟΤΗΤΕΣ ΠΡΟΗΓΟΥΜΕΝΩΝ ΕΒΔΟΜΑΔΩΝ 1Η")

        status, source, confidence = infer_status_from_row(ws, 8, 5, 99)
        self.assertEqual((status, source, confidence), ("pending", "previous_weeks", "unmatched_previous_week_value"))

    def test_infer_orange_day_without_duplicate_is_supervisor_pending(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(9, 2, "ΟΔΟΣ ΔΟΚΙΜΗΣ")
        col = DAY_COLUMNS[0]
        ws.cell(9, col, 3)
        ws.cell(9, col).fill = ORANGE_FILL
        ws.cell(9, col).font = Font(color="FF000000")
        ws.cell(5, col, "ΔΕΥΤΕΡΑ")

        status, source, confidence = infer_status_from_row(ws, 9, col, 3)
        self.assertEqual(status, "pending")
        self.assertEqual(source, "supervisor")
        self.assertEqual(confidence, "pending_day_orange_fill_dark_font")

    def test_infer_orange_day_red_font_is_veltio_pending(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(10, 2, "ΟΔΟΣ ΔΟΚΙΜΗΣ")
        col = DAY_COLUMNS[1]
        ws.cell(10, col, 5)
        ws.cell(10, col).fill = ORANGE_FILL
        ws.cell(10, col).font = Font(color="FFFF0000")
        ws.cell(5, col, "ΤΡΙΤΗ")

        status, source, confidence = infer_status_from_row(ws, 10, col, 5)
        self.assertEqual(status, "pending")
        self.assertEqual(source, "veltio")
        self.assertEqual(confidence, "pending_day_orange_fill_red_font")

    def test_infer_conditional_no_fill_black_font_supervisor_pending(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(11, 2, "ΟΔΟΣ ΔΟΚΙΜΗΣ")
        col = DAY_COLUMNS[2]
        ws.cell(11, col, 8)
        ws.cell(11, col).fill = PatternFill(fill_type="solid", fgColor="00000000")
        ws.cell(11, col).font = Font(color="FF000000")
        ws.cell(5, col, "ΤΕΤΑΡΤΗ")

        status, source, confidence = infer_status_from_row(ws, 11, col, 8)
        self.assertEqual(status, "pending")
        self.assertEqual(source, "supervisor")
        self.assertEqual(confidence, "pending_day_conditional_dark_font")

    def test_summary_duplicate_row_counts_one_collected_on_canonical_day_column(self) -> None:
        """Μετά το dedupe σε επίπεδο εξαγωγής: μία εγγραφή συλλεχθέν στην ημέρα (όχι δύο KPI)."""
        records = [
            DashboardRecord(
                workbook="test.xlsx",
                sheet="Sheet1",
                week_label="week",
                row=8,
                column="F8",
                street="A",
                street_normalized="A",
                period="ΔΕΥΤΕΡΑ",
                day="ΔΕΥΤΕΡΑ",
                point_id=12,
                status="collected",
                source="unknown",
                color="",
                confidence="inferred_duplicate_day_value",
                note="",
            ),
        ]

        summary = build_summary(Path("test.xlsx"), records, [])

        self.assertEqual(summary.collected_by_day, {"ΔΕΥΤΕΡΑ": 1})
        self.assertEqual(summary.resolved_previous_week_records, 0)
        self.assertEqual(summary.collected, 1)

    def test_extract_records_duplicate_two_days_keeps_rightmost_collected(self) -> None:
        import tempfile

        wb = Workbook()
        ws = wb.active
        ws["F2"] = "Εβδομάδα δοκιμής"
        ws.cell(8, 2, "ΟΔΟΣ ΔΟΚΙΜΗΣ")
        c1, c2 = DAY_COLUMNS[0], DAY_COLUMNS[1]
        ws.cell(5, c1, "ΔΕΥΤΕΡΑ")
        ws.cell(5, c2, "ΤΡΙΤΗ")
        ws.cell(8, c1, 99)
        ws.cell(8, c1).fill = ORANGE_FILL
        ws.cell(8, c1).font = Font(color="FF000000")
        ws.cell(8, c2, 99)
        ws.cell(8, c2).fill = GREEN_FILL

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "dup_day.xlsx"
            wb.save(path)
            records, _ = extract_records(path)

        collected = [r for r in records if r.status == "collected"]
        self.assertEqual(len(collected), 1)
        self.assertEqual(collected[0].point_id, 99)
        self.assertEqual(collected[0].day, "ΤΡΙΤΗ")


if __name__ == "__main__":
    unittest.main()
