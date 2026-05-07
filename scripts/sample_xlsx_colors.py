"""One-off helper: print (column, fill_hex, font_descriptor) counts from workbook."""
from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from dashboard_extractor import DATA_COLUMNS, DATA_START_ROW


def fill_rgb(cell) -> str:
    color = cell.fill.fgColor
    if color.type == "rgb" and color.rgb:
        return str(color.rgb).upper()
    return ""


def font_info(cell) -> str:
    fc = cell.font.color
    if fc is None:
        return ""
    if fc.type == "rgb" and fc.rgb:
        return str(fc.rgb).upper()
    if fc.type == "indexed":
        return f"indexed:{fc.indexed}"
    if fc.type == "theme":
        return f"theme:{fc.theme}+{getattr(fc, 'tint', 0)}"
    return str(fc.type)


def main() -> None:
    root = ROOT
    paths = sorted(p for p in root.glob("*.xlsx") if not p.name.startswith("~$"))
    if not paths:
        raise SystemExit(f"No xlsx in {root}")
    wb_path = paths[0]
    print("workbook", wb_path.name)
    ws = load_workbook(wb_path, data_only=False).active
    pairs: Counter[tuple[int, str, str]] = Counter()
    for row in range(DATA_START_ROW, ws.max_row + 1):
        for col in DATA_COLUMNS:
            cell = ws.cell(row, col)
            v = cell.value
            if not isinstance(v, (int, float)) or isinstance(v, bool):
                continue
            pairs[(col, fill_rgb(cell), font_info(cell))] += 1
    for (col, f, g), n in pairs.most_common(50):
        print(col, n, repr(f), repr(g))


if __name__ == "__main__":
    main()
